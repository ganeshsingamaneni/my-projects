import pandas as pd
from itertools import islice
import openpyxl

# path = "/home/ganesh/Documents/Cover_Analysis_from_IDS.xlsx"
# path = "/home/ganesh/Documents/230819.xlsx"
# num = "1"
def fun2(types, dictkeys):
    dict2 = {}
    # print(types,(dictkeys))
    for x in types:
        if x in list(dictkeys):
            pass
        else:
            dict2.update({x:0})
    return dict2


def fun1(l1, l2):
    types = ["BREAK FAST", "LUNCH", "DINNER", "SNACKS"]
    new_dict = {}
    # new_dict={"BREAK FAST": "   ","LUNCH":"   ","DINNER":"   ","SNACKS":"   "}
    try:
        if len(l1) == len(l2):
            pass
    except Exception as e:
        return e
    # print(l1,len(l1))
    # print(l2,len(l2))    
    if len(l1) == 10:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}

        z.update({'Total': l2[5]})
        return z
    elif len(l1) == 10:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[8]: l2[8]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]})
        return z
    elif len(l1) == 16:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]})
        return z
    elif len(l1) == 18:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]})
        z.update({l1[16]: l2[16]})
        return z
    elif len(l1) == 22:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        new_dict.update({l1[13]: l2[17]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]+l2[17]})
        return z
    elif len(l1) == 24:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        new_dict.update({l1[13]: l2[17]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]+l2[17]})
        z.update({l1[22]: l2[22]})
        return z
    elif len(l1) == 28:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        new_dict.update({l1[13]: l2[17]})
        new_dict.update({l1[19]: l2[23]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]+l2[17]+l2[23]})
        return z
    elif len(l1) == 30:
        new_dict.update({l1[0]: l2[0]})
        new_dict.update({l1[1]: l2[5]})
        new_dict.update({l1[7]: l2[11]})
        new_dict.update({l1[13]: l2[17]})
        new_dict.update({l1[19]: l2[23]})
        ff = fun2(types, new_dict.keys())
        z = {**new_dict, **ff}
        z.update({'Total': l2[5]+l2[11]+l2[17]+l2[23]})
        z.update({l1[28]: l2[28]})
        return z
    else:
        pass

def main_func(path):
    wb_obj = openpyxl.load_workbook(path,data_only=True)
    sheet_obj = wb_obj.active
    items_list = ["BAKERY", "MEKONG", "MYSTIQUE LOUNGE",
                "ROOM SERVICE", "SAFFRON SOUL","BANQUET","MYSTIQUE"]
    m_row = sheet_obj.max_row
    column1 = []
    for i in range(5, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=1)
        column1.append(cell_obj.value)
    column2 = []
    for i in range(5, m_row + 1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        column2.append(cell_obj.value)
    # print(column2)    
    ind_list = []
    for x in column1:
        if x in items_list:
            ind = column1.index(x)
            ind_list.append(ind)
    column1_data = [column1[i: j]
                    for i, j in zip([0] + ind_list, ind_list + [None])]
    c1_data = [x for x in column1_data if x != []]
    c1_data = [x for x in c1_data if x != [None]]
    # print(ind_list)
    # print(c1_data)
    for x in c1_data:
        for y in x:
            ii = x.index(y)
            if y is None:
                pass
            elif y.startswith("______________"):
                x[ii] = "    "
            else:
                pass
    column2_data = [column2[i: j]
                    for i, j in zip([0] + ind_list, ind_list + [None])]

    c2_data = [x for x in column2_data if x != []]
    c2_data = [x for x in c2_data if x != [None]]
    for x in c2_data:
        for y in x:
            inn = x.index(y)
            if y == '':
                x[inn] = 0
            elif y is None:
                x[inn] = 0
            else:
                x[inn] = int(y)
    json_list = []
    for a, b in zip(c1_data, c2_data):
        dict_data = fun1(a, b)
        json_list.append(dict_data)
    return json_list    